#what questions do we want to be able to answer?

#1. What weeks in SAT are uncovered? - This can probably be its own function accepting input 'uncovered'
#2. Where is Jordan on date - Done
#3. Who is in SAT on 3/17/20XX - Done


import datetime
from datetime import timedelta, date
import csv
from openpyxl import load_workbook
import openpyxl
import csv
import itertools


class Employee:
    def __init__(self, name1, name2, seniority, base, direct, quota):
        self.name1 = name1 # first name. Class object name should be the employee's last name
        self.name2 = name2 #last name
        self.seniority = seniority #<- datetime object that is the employee's seniority date
        self.base = base #string that is where the employee is based
        self.direct = direct #this should be a true or false for directs vs contractors
        self.quota = quota #<- integer that stores how many weekends are covered

    def __str__(self):
        return self.name2



class station:
    def __init__(self, name, engineer_week, engineer_weekend, shift, weekends):
        self.name = name  # string
        self.engineer_week = engineer_week  # required engineers per week
        self.engineer_weekend = engineer_weekend  # required engineers per weekend
        self.shift = shift  # list of shift type objects
        self.weekends = weekends #self.engineer_week*len(self.shift)  # integer that is the total number of weekends required at this station

    def __str__(self):
        return self.name


class shift:
    def __init__(self,name, dates, engr, sched):
        self.name = name #string that is the week# from excel/csv
        self.dates = dates #list of dates from excel/csv
        self.engr = engr #this is going to be the name of the person covering this shift
        self.sched = sched #7x7, SAL_7x7, 5-2, etc.

    def __str__(self):
        return self.name

#######################################################################################################################
#Employees
Slusky = Employee('Sam','slusky', datetime.datetime(2015, 6, 9), 'CLT', 'y', 0,)
Hagelgans = Employee('Adam','hagelgans', datetime.datetime(2015, 4, 13), 'CLT', 'y', 0)
Tricanowicz = Employee('Mike','tricanowicz', datetime.datetime(2011, 7, 1),'CLT', 'y',  0)
James = Employee('Emily','james', datetime.datetime(2015, 5, 20),'CLT', 'y',  0)
Hogrefe = Employee('Colin','hogrefe', datetime.datetime(2015, 1, 5),'CLT', 'y',  0,)
Porter = Employee('Robb','porter', datetime.datetime(2014, 12, 14),'CLT', 'y',  0)
Bailey = Employee('Zach','bailey', datetime.datetime(2014, 1, 21),'CLT', 'y',  0)
Greer = Employee('Josh','greer', datetime.datetime(2012, 5, 23),'CLT', 'y',  0)
Stedje = Employee('John','stedje', datetime.datetime(2011, 10, 31),'CLT', 'y',  0)
Haines = Employee('Jordan','haines', datetime.datetime(2011, 3, 21),'CLT', 'y',  0)
Cagle = Employee('John','cagle', datetime.datetime(2010, 7,12), 'CLT', 'y', 0)
Malek = Employee('Devin','malek', datetime.datetime(2005, 6, 13),'CLT', 'y', 0)
Monyette = Employee('Tom','monyette', datetime.datetime(2004, 6, 1),'CLT', 'y', 0)
Floren = Employee('Luis','floren', datetime.datetime(1995, 8, 14),'CLT', 'y', 0)

roster = ['slusky', 'hagelgans', 'tricanowicz', 'haines', 'greer']

#Stations
SAT = station('SAT', 1, 1,'',52)
SAL = station('SAL', 2, 2,'', 104)
BFM = station('BFM', 1, 1,'', 52)
GSO = station('GSO', 1, 1, '', 52)
TBD = station('TBD', 1, 1,'', 52)
CLT = station('CLT', 5, 1, '', 52)
VAC = station('VAC', 1, 1, '', 52)

locations = ['CLT', 'SAT', 'BFM', 'GSO', 'JAX', 'SAL', 'VAC']


########################################################################################

#Opening the existing file and then sheet
wb = openpyxl.load_workbook('Sched_master.xlsx')

#Combinations:
# - Station 1, engineer 2 - Done
    # - Right and Wrong Work
# - Station 1, date 2
    #works with clean input of 'DD-mon-YY' which can be changed later
    #date validation in effect
# - Engineer 1, station 2 - Done
    # works
    # probably need to come up with a module that counts how many weeks, maybe store in the class
# - Engineer 1, date 2 - Done
    #-  I think all of this works
# - Date 1, station 2 - Done
    #pretty sure this works
# - Date 1, engineer 2 - Done
    #this one works too

#want to build function to make datetime format easier. Use splitting like reddit function.
########################################################## Below is the module for station first, engr/date second

# The control function takes the mode input of either info or swap. The info function will return a date, engineer,
# or station output. The swap function will make changes to the schedule
def control_func(mode):
    if mode == 'info':
        output = find_info_func(locations, roster, wb)
        return output
    elif mode == 'swap':
        output = swap_func(locations, roster, wb)
        return output


#the find info function works by taking two inputs (2 of date, engineer, or station) and then passes them to the
#appropriate search function which will then return the 3rd piece of data (date, engineer, or station).
def find_info_func(locations, roster, wb):
    data_point1 = '26-APR-2020' #input('station, engineer, or date') #'slusky' #datetime.datetime(2020, 11, 19, 0, 0)
    data_point2 = 'haines' #input('station, engineer, or date') # datetime.datetime(2020, 11, 19, 0, 0)
    if data_point1 in locations:
        station_first(data_point1, data_point2, roster, wb)
    elif data_point1 in roster:
        engr_first(data_point1, data_point2, roster, locations, wb)
    else:
        try:
            date_first(data_point1, data_point2, roster, wb)
            if date_first == None:
                print('NDO')
        except:
            print('invalid input')

#function that is called if the station is the first data point provided
def station_first(data_point1, data_point2, roster, wb):
    print ('station first')
    ws = wb[data_point1] # this is opening the worksheet since station is our primary piece
    #station, engineer -> dates
    row_list = []
    if data_point2 in roster:
        for row in ws.values:
            if data_point2 in row:
                row_list = list(row)
                print (data_point2, 'working', data_point1, row_list) #we'll want to clean this output later
        if not row_list:
            print(data_point2, 'not working in', data_point1, 'that week')
    #station, date -> engineer
    elif data_point2 not in roster:
        validation = 0
        data_point2_fixed = datetime.datetime.strptime(data_point2, '%d-%b-%y')# daay-month-year
        for row in ws.values:
            row_list = list(row)
            if data_point2_fixed in row_list:
                validation = 1
                if row_list[-1]:
                    print(row_list[-1])
                else:
                    print('no coverage')
        if validation != 1:
            print ('invalid date')
    else: print ('done')

#function that is called if the engineer is the first data point provided
def engr_first(data_point1, data_point2, wb):
    print('engr first')
    stations = wb.sheetnames
    #engineer, station -> date
    for i in stations:
        if i == data_point2:
            e_s_validation = 0
            print (i)
            ws = wb[i]
            for row in ws.values:
                row_list = list(row)
                if data_point1 in row_list:
                    n = len(row_list)
                    pretty = [item.strftime('%d-%b-%Y') for item in itertools.islice(row_list, 1,n-1 )]
                    print (data_point1, 'working', pretty)
                    e_s_validation = 1
            if e_s_validation !=1:
                print (data_point1, 'not working', i)
    #engineer, date ->station
    if data_point2 not in stations:
        validation = 0
        data_point2_fixed = datetime.datetime.strptime(data_point2, '%d-%b-%Y')
        for i in stations:
            ws = wb[i]
            for row in ws.values:
                row_list = list(row)
                if data_point2_fixed in row_list:
                    if data_point1 in row_list:
                        print (data_point1, 'in', i)
                        validation = 1
                    else:
                        continue
                        #print (data_point1, 'not at', i, 'on', data_point2_fixed)
        if validation != 1:
            print (data_point1, 'on NDO')

#function that is called if the date is the first data point provided
def date_first(data_point1, data_point2, roster, wb):
    print('date first')
    # date, engineer -> station
    if data_point2 in roster:
        data_point1_fixed = datetime.datetime.strptime(data_point1, '%d-%b-%Y')
        print('date, engineer')
        status = 'NDO'
        stations = wb.sheetnames
        for sheet in stations:
            ws = wb[sheet]
            for row in ws.values:
                if data_point1_fixed in row:
                    if data_point2 in row:
                        status = sheet
                else:
                    continue
        print(data_point2, 'in', status)
    #date, station -> engineer
    elif data_point2 not in roster:
        print('date, station')
        data_point1_fixed = datetime.datetime.strptime(data_point1, '%d-%b-%Y')
        ws = wb[data_point2]
        for row in ws.values:
            row_list = list(row)
            if data_point1_fixed in row_list: #this is where we're not making it
                engr = row_list[-1]
                if engr:
                    print (engr, 'working', row_list[0])
                if not engr:
                    print (row_list[0], 'uncovered')


#The swap function requires a date, station, and engineer input, as well as an 'in' or .
# The 'in' input will have the function insert the given engineer at the given station on the given date
# the 'out' input will have the function remove the given engineer from the given station on the given date
def swap_func(locations, roster, wb): #there will hae to be a prompt for swap in or swap out
    # right now this function works for a single date, will want to allow it to work on a series of dates
    station = 'CLT'
    date = '26-NOV-2020'
    date_fixed = datetime.datetime.strptime(date, '%d-%b-%Y')# daay-month-year
    new_engr = 'greer'
    direction = 'out'
    ws = wb[station]
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == date_fixed: #checking the value of the cell not the cell object for the date
                existing_engr_cell = [row[-1].coordinate]
                if direction == 'in':
                    ws[existing_engr_cell[0]] = new_engr
                elif direction == 'out':
                    ws[existing_engr_cell[0]] = None
    wb.save("C:\\Users\\sam\\PycharmProjects\\work_calendar\\Sched_master.xlsx")
    print ('swap made')




#########################################################



mode = input('mode:')

control_func(mode)
